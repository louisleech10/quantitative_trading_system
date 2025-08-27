#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
K線數據下載驗證工具
用於下載指定日期的BTCUSDT原始數據，供手動驗證計算結果

使用方法：
1. 修改 VERIFICATION_CASES 中的案例日期
2. 運行腳本：python data_verification_downloader.py
3. 檢查生成的Excel檔案進行手動驗證
"""

import sys
import os
import pandas as pd
import requests
from datetime import datetime, timedelta
from pathlib import Path
import time

# 添加項目路徑
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)
sys.path.append(parent_dir)

# ===== 使用者配置區域 START =====

# 要驗證的案例日期列表 - 您可以修改這些日期
VERIFICATION_CASES = [
    "2024-02-09",  # 從CSV中挑選的第1個案例
    "2024-02-12",  # 從CSV中挑選的第2個案例
    "2024-02-14",  # 從CSV中挑選的第3個案例
    "2024-03-11",  # 從CSV中挑選的第4個案例
    "2024-07-15",  # 從CSV中挑選的第5個案例
]

# 配置參數
CONFIG = {
    'symbol': 'BTCUSDT',
    'interval': '1d',           # K線時間週期
    'lookback_days': 25,        # 每個案例向前獲取的天數
    'forward_days': 15,         # 每個案例向後獲取的天數
    'output_dir': 'verification_data',  # 輸出目錄
}

# ===== 使用者配置區域 END =====

class BinanceDataDownloader:
    """幣安數據下載器"""
    
    def __init__(self):
        self.base_url = "https://api.binance.com/api/v3/klines"
        self.session = requests.Session()
        
    def get_klines(self, symbol, interval, start_time, end_time, limit=1000):
        """獲取K線數據"""
        params = {
            'symbol': symbol,
            'interval': interval,
            'startTime': int(start_time.timestamp() * 1000),
            'endTime': int(end_time.timestamp() * 1000),
            'limit': limit
        }
        
        try:
            response = self.session.get(self.base_url, params=params)
            response.raise_for_status()
            
            data = response.json()
            
            # 轉換為DataFrame
            columns = [
                'open_time', 'open', 'high', 'low', 'close', 'volume',
                'close_time', 'quote_asset_volume', 'number_of_trades',
                'taker_buy_base_asset_volume', 'taker_buy_quote_asset_volume', 'ignore'
            ]
            
            df = pd.DataFrame(data, columns=columns)
            
            # 轉換數據類型
            numeric_columns = ['open', 'high', 'low', 'close', 'volume', 
                             'quote_asset_volume', 'taker_buy_base_asset_volume', 
                             'taker_buy_quote_asset_volume']
            
            for col in numeric_columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')
            
            # 轉換時間
            df['datetime'] = pd.to_datetime(df['open_time'], unit='ms')
            df['date'] = df['datetime'].dt.date
            
            # 確保datetime列是正確的datetime類型
            df['datetime'] = pd.to_datetime(df['datetime'])
            
            # 計算基本指標
            df['price_change_pct'] = ((df['close'] - df['open']) / df['open'] * 100).round(4)
            df['high_low_range_pct'] = ((df['high'] - df['low']) / df['open'] * 100).round(4)
            
            return df[['datetime', 'date', 'open', 'high', 'low', 'close', 'volume', 
                      'price_change_pct', 'high_low_range_pct', 'number_of_trades']]
            
        except Exception as e:
            print(f"獲取K線數據時出錯: {str(e)}")
            return None
    
    def download_case_data(self, target_date, lookback_days, forward_days):
        """下載指定案例的數據"""
        # 計算時間範圍
        target_dt = datetime.strptime(target_date, '%Y-%m-%d')
        start_time = target_dt - timedelta(days=lookback_days)
        end_time = target_dt + timedelta(days=forward_days)
        
        print(f"下載 {target_date} 的數據...")
        print(f"  時間範圍: {start_time.date()} 到 {end_time.date()}")
        
        # 獲取數據
        df = self.get_klines(
            symbol=CONFIG['symbol'],
            interval=CONFIG['interval'],
            start_time=start_time,
            end_time=end_time
        )
        
        if df is not None:
            # 找到目標日期的索引
            target_row = df[df['date'] == target_dt.date()]
            if not target_row.empty:
                target_index = target_row.index[0]
                
                # 添加相對位置列
                df['days_from_target'] = (pd.to_datetime(df['date']) - target_dt).dt.days
                df['is_target_date'] = df['date'] == target_dt.date()
                
                # 計算未來收益率（向前1-10天）
                for i in range(1, 11):
                    if target_index + i < len(df):
                        future_close = df.iloc[target_index + i]['close']
                        target_close = df.iloc[target_index]['close']
                        df.loc[target_index, f'future_{i}d_return_pct'] = round(
                            ((future_close - target_close) / target_close * 100), 4
                        )
                
                print(f"  成功獲取 {len(df)} 條數據")
                return df
            else:
                print(f"  警告: 未找到目標日期 {target_date} 的數據")
                return df
        else:
            print(f"  失敗: 無法獲取數據")
            return None

def create_verification_excel():
    """創建驗證用的Excel檔案"""
    
    # 創建輸出目錄
    output_dir = Path(CONFIG['output_dir'])
    output_dir.mkdir(exist_ok=True)
    
    # 初始化下載器
    downloader = BinanceDataDownloader()
    
    # 創建Excel writer
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    excel_path = output_dir / f'btcusdt_verification_data_{timestamp}.xlsx'
    
    successful_cases = 0  # 追蹤成功處理的案例數
    
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        
        # 為每個案例創建工作表
        for i, case_date in enumerate(VERIFICATION_CASES, 1):
            try:
                print(f"\n[{i}/{len(VERIFICATION_CASES)}] 處理案例: {case_date}")
                
                # 下載數據
                df = downloader.download_case_data(
                    target_date=case_date,
                    lookback_days=CONFIG['lookback_days'],
                    forward_days=CONFIG['forward_days']
                )
                
                if df is not None:
                    # 創建工作表名稱
                    sheet_name = f'Case_{i}_{case_date.replace("-", "")}'
                    
                    # 寫入工作表
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    successful_cases += 1  # 增加成功計數
                    
                    # 獲取工作表以設置格式
                    worksheet = writer.sheets[sheet_name]
                    
                    # 設置列寬
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 20)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    # 高亮目標日期行
                    from openpyxl.styles import PatternFill
                    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    
                    for row in worksheet.iter_rows():
                        if row[1].value and str(row[1].value)[:10] == case_date:  # 檢查date列
                            for cell in row:
                                cell.fill = yellow_fill
                
                # 避免API限制
                time.sleep(0.1)
                
            except Exception as e:
                print(f"處理案例 {case_date} 時出錯: {str(e)}")
                continue
        
        # 檢查是否有成功的案例
        if successful_cases == 0:
            print("\n警告: 沒有成功處理任何案例，創建空白工作表以避免錯誤")
            # 創建一個空白工作表
            empty_df = pd.DataFrame({'說明': ['所有案例處理失敗，請檢查網路連接或日期格式']})
            empty_df.to_excel(writer, sheet_name='Error_Log', index=False)
    
    return excel_path

def create_calculation_reference():
    """創建計算公式參考檔案"""
    
    reference_content = """
# BTCUSDT 數據驗證計算參考

## 基本計算公式

### 1. 價格變化百分比 (price_change)
price_change = (close - open) / open * 100

### 2. 未來收益率 (future_Xbar_return)  
future_1bar_return = (次日close - 當日close) / 當日close * 100
future_2bar_return = (後2日close - 當日close) / 當日close * 100
...以此類推

### 3. 收盤強度 (closing_strength)
closing_strength = (close - low) / (high - low)
範圍: 0-1，數值越高表示收盤價越接近最高價

### 4. 價格位置 (price_position)  
需要過去20天的數據計算
price_position = (當日close - 20日最低) / (20日最高 - 20日最低)

### 5. 成交量倍數 (volume_multiplier)
需要過去20天的平均成交量
volume_multiplier = 當日volume / 20日平均volume

## 驗證重點

1. **目標日期確認**: Excel中黃色高亮的行就是您要驗證的案例日期
2. **基本計算**: 先驗證 price_change 是否 >= 3%
3. **未來收益**: 檢查 future_1d_return_pct 等欄位是否與系統計算一致
4. **數據完整性**: 確認目標日期前後都有足夠的數據用於計算

## 手動驗證步驟

1. 在Excel中找到黃色高亮的目標日期行
2. 用計算器驗證 price_change_pct 欄位: (收盤價-開盤價)/開盤價*100
3. 驗證未來收益率: 找到後1天、後2天的收盤價進行計算
4. 對比系統計算的結果，標記差異

如有計算差異，請記錄下來以便修正系統邏輯。
"""
    
    # 保存參考檔案
    output_dir = Path(CONFIG['output_dir'])
    ref_path = output_dir / 'calculation_reference.md'
    
    with open(ref_path, 'w', encoding='utf-8') as f:
        f.write(reference_content)
    
    return ref_path

def main():
    """主函數"""
    print("BTCUSDT 數據驗證工具")
    print("=" * 50)
    
    print(f"配置信息:")
    print(f"  交易對: {CONFIG['symbol']}")
    print(f"  時間週期: {CONFIG['interval']}")
    print(f"  向前天數: {CONFIG['lookback_days']}")
    print(f"  向後天數: {CONFIG['forward_days']}")
    print(f"  驗證案例: {len(VERIFICATION_CASES)}個")
    
    try:
        # 創建Excel驗證檔案
        print("\n開始下載驗證數據...")
        excel_path = create_verification_excel()
        print(f"\nExcel檔案已創建: {excel_path}")
        
        # 創建計算參考
        ref_path = create_calculation_reference()
        print(f"計算參考檔案已創建: {ref_path}")
        
        print("\n" + "=" * 50)
        print("下載完成!")
        print("\n下一步:")
        print(f"1. 打開Excel檔案: {excel_path}")
        print(f"2. 參考計算說明: {ref_path}")
        print("3. 黃色高亮的行是您要驗證的目標日期")
        print("4. 手動驗證基本計算公式是否正確")
        print("5. 將驗證結果與系統計算對比")
        
    except Exception as e:
        print(f"\n執行過程中出錯: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()